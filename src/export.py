from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
PCT_FMT = "0.00%"
NUM_FMT = "#,##0.00"


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data(ws, row_count, col_count, pct_cols=None):
    pct_cols = pct_cols or []
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if col in pct_cols:
                cell.number_format = PCT_FMT
            elif isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT


def _auto_width(ws, col_count):
    for col in range(1, col_count + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = row[0]
            if val is not None:
                # 中文字符按2倍宽度计算
                s = str(val)
                length = sum(2 if ord(c) > 127 else 1 for c in s)
                if length > max_len:
                    max_len = length
        ws.column_dimensions[letter].width = min(max_len + 4, 40)


def write_excel(reports: dict[str, "pd.DataFrame"], output_path: str | Path):
    """
    将报表字典写出为格式化的 Excel 文件
    reports: {sheet_name: DataFrame}
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in reports.items():
        if df.empty:
            continue

        ws = wb.create_sheet(title=sheet_name)

        # 写表头
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写数据（替换 NaN/NA 为 None）
        df = df.where(pd.notna(df), None)
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header)
                val = None if pd.isna(val) else val
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 百分比列识别（包含"率"或"比"的列）
        pct_cols = set()
        for col_idx, header in enumerate(headers, 1):
            if any(kw in str(header) for kw in ["率", "比", "占比"]):
                pct_cols.add(col_idx)

        col_count = len(headers)
        row_count = ws.max_row

        _style_header(ws, col_count)
        _style_data(ws, row_count, col_count, pct_cols)
        _auto_width(ws, col_count)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        if row_count > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{row_count}"

    wb.save(output_path)
    return output_path
