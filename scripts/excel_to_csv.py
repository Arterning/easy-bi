import duckdb
import sys
from pathlib import Path
from openpyxl import load_workbook


def get_sheet_names(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    return wb.sheetnames


def excel_to_csv(excel_path: str, output_dir: str | None = None) -> None:
    src = Path(excel_path).resolve(strict=True)

    if output_dir:
        out = Path(output_dir).resolve()
    else:
        out = src.parent / src.stem
    out.mkdir(parents=True, exist_ok=True)

    sheet_names = get_sheet_names(str(src))
    if not sheet_names:
        print(f"No sheets found in {src}", file=sys.stderr)
        sys.exit(1)

    con = duckdb.connect()
    try:
        con.install_extension("excel")
        con.load_extension("excel")
    except Exception as e:
        print(f"Failed to load duckdb excel extension: {e}", file=sys.stderr)
        sys.exit(1)

    for sheet in sheet_names:
        csv_path = out / f"{sheet}.csv"
        csv_esc = str(csv_path).replace("'", "''")
        path_esc = str(src).replace("'", "''")
        sheet_esc = sheet.replace("'", "''")
        sql = (
            f"COPY (SELECT * FROM read_xlsx('{path_esc}', sheet='{sheet_esc}', all_varchar=true)) "
            f"TO '{csv_esc}' (FORMAT CSV, HEADER true)"
        )
        con.execute(sql)
        print(f"Exported: {csv_path}")

    con.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        name = Path(__file__).name
        print(f"Usage: python {name} <excel_path> [output_dir]")
        print(f"Example: python {name} data.xlsx output_csvs")
        sys.exit(1)
    excel_to_csv(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
